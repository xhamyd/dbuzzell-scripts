from datetime import datetime
from pathlib import Path
import os
import re
from typing import Optional

from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook
import requests
import yaml

# Load environmental variables from ".env"
load_dotenv()

GEOCODIO_API_KEY = os.environ["GEOCODIO_API_KEY"]
LEGISLATORS_YAML_URL = "https://unitedstates.github.io/congress-legislators/legislators-current.yaml"
COORDINATES_XLSX = Path(f"./{os.environ['COORDINATES_XLSX']}")
LAT_LON_REGEX = r'^(?P<lat>[\d\.-]+)\u00B0?,\s*(?P<lon>[\d\.-]+)\u00B0?$'
OUTPUT_FILE = Path(f"./output.xlsx")


def get_coordinates() -> tuple:
    workbook = load_workbook(COORDINATES_XLSX)
    sheet = workbook.active
    if sheet is None:
        raise RuntimeError(f"Failed to load worksheet {COORDINATES_XLSX}")

    data = [row for row in sheet.iter_rows(min_row=2, max_row=None, min_col=2, max_col=4, values_only=True)]
    assert data[0] == ('Project Number', 'PROJECT NAME', 'Geo Coordinates')  # assuming a fixed format for now
    assert data[1] == (None, 'TOTAL', '-')  # extra info that we don't need

    return tuple([row_i] + list(row[1:]) for row_i, row in enumerate(data[2:], start=1))


def is_current_term(term):
    """Return True if the term is current (end date in future or missing)."""
    end_date = term.get('end')
    if end_date is None:
        return True

    end_dt = datetime.strptime(end_date, '%Y-%m-%d')
    return end_dt > datetime.now()


def get_legislators_yaml():
    """Fetch and parse the YAML legislators file, returning current House members."""
    response = requests.get(LEGISLATORS_YAML_URL)
    response.raise_for_status()
    data = yaml.safe_load(response.text)

    legislators = dict()
    for leg in data:
        # Out of all legislator terms, find all current Representatives
        rep_terms = [t for t in leg['terms']
                     if t['type'] == 'rep' and is_current_term(t)]
        rep_terms.sort(key=lambda t: t['start'], reverse=True)

        # Use the latest current term
        if not rep_terms:
            continue
        term = rep_terms[0]

        legislators[f"{term['state']}-{term['district']}"] = dict(name=leg['name'], party=term['party'])

    return legislators


def get_district(lat_lon: tuple[Optional[float], ...]) -> dict:
    """Get congressional district using Geocodio"""
    res = dict(state=None,
               country=None,
               district=None)
    if None in lat_lon:
        # One or both coordinates are missing, skip reverse lookup
        return res

    url = "https://api.geocod.io/v1.8/reverse"
    params = dict(q=f"{lat_lon[0]},{lat_lon[1]}",
                  api_key=GEOCODIO_API_KEY,
                  fields='cd')

    # params_str = "&".join(list(f"{k}={v}" for k, v in params.items()))
    # print(f"{url}?{params_str}")

    response = requests.get(url, params=params)
    data = response.json()
    if response.status_code != 200:
        print(f"Error {response.status_code}: {data['error']}")
        return res

    results = data['results'][0]
    res['state'] = results['address_components']['state']
    res['country'] = results['address_components']['country']
    if res['country'] != "CA":
        # Skip canadian addresses
        cd_info = results['fields']['congressional_districts'][0]
        res['district'] = cd_info['district_number']
    # else:
    #     breakpoint()
    
    return res


def ordinal(n: Optional[int]) -> str:
    if n is None:
        return "#N/A"
    else:
        # Source: https://codegolf.stackexchange.com/questions/4707/outputting-ordinal-numbers-1st-2nd-3rd#answer-4712
        return "%d%s" % (n, "tsnrhtdd"[(n//10%10!=1)*(n%10<4)*n%10::4])


def match_legislators_with_coords(proj_info: dict, lat_lon_str: str, legislators: dict) -> dict:
    legis_res = dict(proj_num=proj_info['num'],
                     proj_name=proj_info['name'])
    print(f"{proj_info['num']}: {proj_info['name']}")

    if lat_lon_str is None:
        print("*** No coordinates provided!")
        lat_lon = (None, None)
    elif (lat_lon_match := re.match(LAT_LON_REGEX, lat_lon_str)):
        lat_lon = (float(lat_lon_match.group("lat")),
                   float(lat_lon_match.group("lon")))
    # else:
    #     breakpoint()
    legis_res['lat_lon'] = lat_lon
    print(f"Coordinates: {lat_lon}")

    district_info = get_district(lat_lon)
    state_str = "??" if district_info['state'] is None else district_info['state']
    if district_info['country'] == "CA":
        # Use Canada as the District Number
        district_str = district_info['country']
    elif district_info['district'] is None:
        # Missing district number
        district_str = "??"
    else:
        # Space formatting for even-align
        district_str = f"{district_info['district']:>02d}"
    district = f"{state_str}-{district_str}"
    legis_res['district'] = district
    print(f"District: {district}")

    representative = dict(first_name=None, last_name=None, party=None)
    if None not in (district_info['state'], district_info['district']):
        # District lookup succeeded, find the legislator
        key = f"{district_info['state']}-{district_info['district']}"  # STRICT KEY FORMATTING!
        if key in legislators:
            rep = legislators[key]
            representative['first_name'] = rep['name']['first']
            representative['last_name'] = rep['name']['last']
            representative['party'] = rep['party']
    legis_res['rep'] = representative

    if representative['first_name'] is None and representative['last_name'] is None:
        rep_name_str = "#N/A"
    else:
        rep_name_str = f"{legis_res['rep']['first_name']} {legis_res['rep']['last_name']}"
    print(f"Representative: {rep_name_str} ({legis_res['rep']['party']})\n")

    print()  # newline
    return legis_res


def write_output_xlsx(proj_coords: list[dict]):
    wb = Workbook()
    ws = wb.active  # Sheet1
    assert ws is not None

    # Write the headers
    headers = ['Project Number', 'Latitude', 'Longitude', 'Congressional District', 'Representative Name', 'Representative Party']
    for col_i, head in enumerate(headers, start=1):
        ws.cell(row=1, column=col_i).value = head

    def get_col_ind(head: str):
        return headers.index(head) + 1  # Excel is one-indexed

    # Write the project rows
    for row_i, proj_info in enumerate(proj_coords, start=2):
        ws.cell(row=row_i, column=get_col_ind('Project Number')).value = proj_info['proj_num']
        ws.cell(row=row_i, column=get_col_ind('Latitude')).value = proj_info['lat_lon'][0]
        ws.cell(row=row_i, column=get_col_ind('Longitude')).value = proj_info['lat_lon'][1]
        ws.cell(row=row_i, column=get_col_ind('Congressional District')).value = proj_info['district']
        if None in (proj_info['rep']['first_name'], proj_info['rep']['last_name']):
            rep_name = "#N/A"
        else:
            rep_name = f"{proj_info['rep']['first_name']} {proj_info['rep']['last_name']}"
        ws.cell(row=row_i, column=get_col_ind('Representative Name')).value = rep_name
        ws.cell(row=row_i, column=get_col_ind('Representative Party')).value = proj_info['rep']['party']

    wb.save(OUTPUT_FILE)


def main():
    print("Getting legislators...")
    legislators = get_legislators_yaml()
    print("Loading coordinates...")
    coordinates = get_coordinates()
    print("Done!\n")

    output_info = list()
    for proj_num, proj_name, lat_lon_str in coordinates:
        if int(proj_num) > 245:
            # Reached end of list
            break

        proj_info = dict(num=proj_num, name=proj_name)
        legis_coords = match_legislators_with_coords(proj_info, lat_lon_str, legislators)
        output_info.append(legis_coords)

    write_output_xlsx(output_info)


if __name__ == "__main__":
    main()
